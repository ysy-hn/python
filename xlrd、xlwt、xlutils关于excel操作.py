# pands操作excel更容易
#
# xlrd:对excel内容读取。
# xlwt:对excel内容写入。
# 使用pip安装，注：xlrd 1.2.0版本可对xlsx操作，高版本只能对xls操作；因为加密软件导致无法打开
# 获取工作簿对象：book = xlrd.open_workbook('excel文件名称')
# 获取所有工作表名称：names = book.sheet_names()，结果为列表
# 根据索引获取工作表对象：sheet = book.sheet_by_index(i)
# 根据名称获取工作表对象：sheet = book.sheet_by_name('工作表名称')
# 获取工作表行数：rows = sheet.nrows
# 获取工作表列数：cols = sheet.ncols
# 获取工作表某一行的内容：row = sheet.row_values(i) ，结果为列表   【sheet.row(i)，列表】
# 获取工作表某一列的内容：col = sheet.col_values(i)  结果为列表   【sheet.col(i)，列表】
# 获取工作表某一单元格的内容：cell = sheet.cell_value(m,n)、 sheet.cell(m,n).value、sheet.row(m)[n].value，
# sheet.col(n)[m].value，结果为字符串或数值    【sheet.cell(0,0)，xlrd.sheet.Cell对象】
#
#
# 创建工作簿：book = xlwt.Workbook()，如果写入中文为乱码，可添加参数encoding = 'utf-8'
# 创建工作表：sheet = book.add_sheet('Sheet1')
# 向单元格写入内容：sheet.write(m,n,'内容1')、sheet.write(x,y,'内容2')
# 保存工作簿：book.save('excel文件名称')，默认保存在py文件相同路径下，如果该路径下有相同文件，
# 会被新创建的文件覆盖，即xlwt不能修改文件。
#
# openpyxl模块可实现对excel文件的读、写和修改，只能处理xlsx文件，不能处理xls文件，先安装该模块;
# 对于openpyxl，行数和列数都从1开始，单元格的行和列也从1开始。
# 获取工作簿对象：book = openpyxl.load_workbook('excel文件名称')
# 获取所有工作表名称：names = book.sheetnames
# 获取工作表对象：sheet1 = book.worksheets[n]、sheet2 = book['工作表名称']、sheet3 = book[book.sheetnames[n]]
# 获取工作表名称：title = sheet1.title
# 获取工作表行数：rows = sheet1.max_row
# 获取工作表列数：cols = sheet1.max_column
# 获取某一单元格内容：cell = sheet.cell(1,2).value、sheet['单元格'].value例如sheet['B1'].value
# 假设有一fruit2.xlsx，除后缀名其他与上述fruit.xls完全一样
#
# 创建工作簿：book = openpyxl.Workbook()，如果写入中文为乱码，可添加参数encoding = 'utf-8'
# 创建工作表：sheet = book.create_sheet('工作表名称',0)，0表示创建的工作表在工作薄最前面
# 向单元格写入内容：sheet.cell(m,n,'内容1')、sheet.cell(x,y,'内容2')
# 保存工作簿：book.save('excel文件名称')，默认保存在py文件相同路径下，如果该路径下有相同文件，会被新创建的文件覆盖。
#
# sheet.insert_rows(m)和sheet.insert_cols(n)分别表示在第m行、第n列前面插入行、列
# sheet.delete_rows(m)和sheet.delete_cols(n)分别表示删除第m行、第n列
#
# python之xlutils的Copy模块,可以复制excel

import xlrd
from xlutils.copy import copy
workbook = xlrd.open_workbook('调试记录0825.xlsx')  # 打开工作簿
new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
new_workbook.save("调试记录0825-2.xlsx")  # 保存工作簿


import openpyxl
book = openpyxl.load_workbook('调试记录0825.xlsx')
names = book.sheetnames
print(names)
sheet = book.worksheets[0]
print(sheet)
sheet2 = book['平整度']
sheet3 = book[book.sheetnames[0]]
print(sheet2, sheet3)
print(sheet3.title)
rows1 = sheet.max_row
cols1 = sheet.max_column
print(rows1, cols1)
cell1 = sheet.cell(9,1)
cell2 = sheet['a10'].value
print(cell1,cell2)
rows2 = sheet.rows
cols2 = sheet.columns
print(rows2, cols2)
for i in list(sheet.rows)[1]:
    print(i.value,end='  ')
print()
for i in list(sheet.columns)[0]:
    print(i.value,end='  ')


import xlrd
# workbook = xlrd.open_workbook('调试记录0825.xlsx')  # 打开excel文件
# sheet_all = workbook.sheet_names()  # 获取所有的sheet名称
# sheet_0 = sheet_all[0]
# sheet_1 = sheet_all[1]
# print(sheet_all, sheet_0, sheet_1)
#
# sheet0 = workbook.sheet_by_index(0)       # 根据索引获取sheet内容，同时获取sheet名称、行数、列数
# sheet1 = workbook.sheet_by_name('平整度')  # 根据名称获取sheet内容，同时获取sheet名称、行数、列数
# print(sheet0.name, sheet0.nrows, sheet0.ncols)  # nrows：行数，ncols：列数
# print(sheet1.name, sheet1.nrows, sheet1.ncols)
#
# rows = sheet0.row_values(9)  # 获取想要行数的内容
# cols = sheet0.col_values(1)  # 获取想要列数的内容
# print(rows)
# print(cols)
#
# print(sheet0.cell(9,0))  # 获取单元格，注释类型
# print(sheet0.cell_value(9,0))  # 只获取单元值
# print(sheet0.row(9)[0])  # 通过行和索引获取值，注释类型
# print(sheet0.cell(9,0).ctype)  # 1：字符串类型；2：数字类型；3：时间类型
# print(sheet0.cell(5,1).ctype)

# 处理date格式时，可以先判断是否为3类型，再进行格式转换操作
# from datetime import datetime, date
# workbook = xlrd.open_workbook('调试记录0825.xlsx')  # 打开excel文件
# sheet = workbook.sheet_by_index(0)
# if sheet.cell(row,col).ctype == 3:
#     date_value = xlrd.xldate_as_tuple(sheet.cell_value(row,col), workbook.datemode)
#     date_tmp = date(*date_value[:3]).strftime('%Y%m%d')

# 获取合并单元格的内容
# 1、读取合并单元格的第一个单元格索引
# 2、读取合并单元格的最大索引，会以列表的形式显示，标红的为合并单元格
# 3、使用merged_cells方法处理，读取合并的第一个cell，错了为空
workbook = xlrd.open_workbook('调试记录0825.xlsx')
sheet = workbook.sheet_by_index(0)
print(sheet.merged_cells)
# merged_cells输出:获取返回的row和col低位的索引即可
merge = []
for (rlow,rhigh,clow,chigh) in sheet.merged_cells:
    merge.append([rlow,clow])
for index in merge:
    print((sheet.cell_value(index[0],index[1])))


# xlwt的使用
# write_merge(x, x + m, y, w + n, string, sytle):x表示行，y表示列，
# m表示跨行个数，n表示跨列个数，string表示要写入的单元格内容，style表示单元格样式。其中，x，y，w，h，都是以0开始计算的
"""
设置单元格样式
"""
import xlwt
def set_style(name, height, bold=False):
    style = xlwt.XFStyle()  # 初始化样式
    font = xlwt.Font()  # 为样式创建字体
    font.name = name  # 'Times New Roman'
    font.blod = bold
    font.colour_index = 4
    font.height = height

    borders = xlwt.Borders()
    borders.left=6
    borders.right=6
    borders.top=6
    borders.bottom = 6

    style.font = font
    style.borders = borders

    return style

# 写excel
def write_excel():
    f = xlwt.Workbook()  # 创建工作簿
    """
    创建第一个sheet
    """
    sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok = True)  # 创建sheet
    row0 = [u'业务', u'状态', u'北京', u'上海', u'广州', u'深圳', u'状态小计', u'合计']
    column0 = [u'机票', u'船票', u'火车票', u'汽车票', u'其它']
    status = [u'预订', u'出票', u'退票', u'业务小计']

    # 生成第一行
    for i in range(0, len(row0)):
        sheet1.write(0, i, row0[i], set_style('Times New Roman', 220, True))

        # 生成第一列和最后一列(合并4行)
    i, j = 1, 0
    while i < 4*len(column0) and j < len(column0):
        sheet1.write_merge(i, i + 3, 0, 0, column0[j], set_style('Arial', 220, True))  # 第一列
        sheet1.write_merge(i, i + 3, 7, 7)  # 最后一列"合计"
        i += 4
        j += 1

    sheet1.write_merge(21, 21, 0, 1, u'合计', set_style('Times New Roman', 220, True))

    # 生成第二列
    i = 0
    while i < 4*len(column0):
        for j in range(0, len(status)):
            sheet1.write(j + i + 1, 1, status[j])
        i += 4

    f.save('demo1.xlsx')  # 保存文件

    '''
      创建第二个sheet:
        sheet2
      '''
    sheet2 = f.add_sheet(u'sheet2', cell_overwrite_ok = True)  # 创建sheet2
    row0 = [u'姓名', u'年龄', u'出生日期', u'爱好', u'关系']
    column0 = [u'小杰', u'小胖', u'小明', u'大神', u'大仙', u'小敏', u'无名']

    # 生成第一行
    for i in range(0, len(row0)):
        sheet2.write(0, i, row0[i], set_style('Times New Roman', 220, True))

    # 生成第一列
    for i in range(0, len(column0)):
        sheet2.write(i + 1, 0, column0[i], set_style('Times New Roman', 220))

    sheet2.write(1, 2, '1991/11/11')
    sheet2.write_merge(7, 7, 2, 4, u'暂无')  # 合并列单元格
    sheet2.write_merge(1, 2, 4, 4, u'好朋友')  # 合并行单元格

    f.save('demo1.xlsx')  # 保存文件

if __name__ == '__main__':
    # generate_workbook()
    # read_excel()
    write_excel()
